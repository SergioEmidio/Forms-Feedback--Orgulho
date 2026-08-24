# app/utils/planilha.py
# Duas direções:
#   1. gerar_planilha_respostas()  -> BANCO vira arquivo .xlsx (exportar/baixar)
#   2. ler_planilha_para_respostas() -> arquivo .xlsx/.csv vira dados prontos
#      pro banco (importar/upload em massa)

import io
from typing import List, Dict

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLUNAS_EXIBICAO = {
    "id": "ID",
    "nome_completo": "Nome completo",
    "nascimento": "Nascimento",
    "tipo_participante": "Tipo de participante",
    "outro_detalhe": "Detalhe (Outro)",
    "nota": "Nota",
    "opiniao": "Opinião",
    "enviado_em": "Enviado em",
}

# Mapa inverso: nome da coluna como aparece na planilha -> nome interno do campo.
# Usado ao IMPORTAR, pra aceitar tanto "Nome completo" quanto "nome_completo"
# quanto "nome-completo" na primeira linha da planilha.
COLUNAS_ACEITAS_NA_IMPORTACAO = {
    "id": "id",
    "nome completo": "nome_completo",
    "nome_completo": "nome_completo",
    "nome-completo": "nome_completo",
    "nascimento": "nascimento",
    "tipo de participante": "tipo_participante",
    "tipo_participante": "tipo_participante",
    "tipo-participante": "tipo_participante",
    "detalhe (outro)": "outro_detalhe",
    "outro_detalhe": "outro_detalhe",
    "outro-detalhe": "outro_detalhe",
    "nota": "nota",
    "opinião": "opiniao",
    "opiniao": "opiniao",
    "enviado em": "enviado_em",
    "enviado_em": "enviado_em",
    "enviado-em": "enviado_em",
    "consentimento": "consentimento",
}


# ============================================================
# EXPORTAR: banco -> arquivo .xlsx
# ============================================================
def gerar_planilha_respostas(respostas: List[Dict]) -> io.BytesIO:
    """Recebe uma lista de dicionários (uma resposta por item) e devolve
    um arquivo .xlsx pronto, já formatado, em memória (BytesIO)."""

    df = pd.DataFrame(respostas)

    colunas_presentes = [c for c in COLUNAS_EXIBICAO if c in df.columns]
    df = df[colunas_presentes] if colunas_presentes else df
    df = df.rename(columns=COLUNAS_EXIBICAO)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Respostas")
        planilha = writer.sheets["Respostas"]

        preenchimento_cabecalho = PatternFill(start_color="0D0D18", end_color="0D0D18", fill_type="solid")
        fonte_cabecalho = Font(color="00FFF9", bold=True)

        for indice_coluna, titulo in enumerate(df.columns, start=1):
            celula = planilha.cell(row=1, column=indice_coluna)
            celula.fill = preenchimento_cabecalho
            celula.font = fonte_cabecalho
            celula.alignment = Alignment(horizontal="center")

            valores_coluna = df.iloc[:, indice_coluna - 1].astype(str)
            maior_largura = max([len(str(titulo))] + [len(v) for v in valores_coluna])
            letra_coluna = get_column_letter(indice_coluna)
            planilha.column_dimensions[letra_coluna].width = min(maior_largura + 4, 50)

        planilha.freeze_panes = "A2"

    buffer.seek(0)
    return buffer


# ============================================================
# IMPORTAR: arquivo .xlsx/.csv -> lista de dicionários prontos pro banco
# ============================================================
def ler_planilha_para_respostas(conteudo_arquivo: bytes, nome_arquivo: str) -> List[Dict]:
    """Lê um arquivo Excel (.xlsx) ou CSV enviado por upload e devolve uma
    lista de dicionários com os nomes de campo internos (nome_completo,
    tipo_participante, etc.) — prontos pra validar com RespostaCreate
    e inserir no banco.

    Não insere nada sozinho: quem chama essa função decide o que fazer
    com o resultado (permite validar linha a linha antes de gravar).
    """
    buffer = io.BytesIO(conteudo_arquivo)

    if nome_arquivo.lower().endswith(".csv"):
        df = pd.read_csv(buffer)
    else:
        df = pd.read_excel(buffer)

    # Normaliza os nomes de coluna (minúsculo, sem espaço nas pontas) antes
    # de traduzir pelo mapa — assim aceita variações de maiúscula/espaço.
    df.columns = [str(c).strip().lower() for c in df.columns]

    colunas_traduzidas = {}
    for coluna_original in df.columns:
        campo_interno = COLUNAS_ACEITAS_NA_IMPORTACAO.get(coluna_original)
        if campo_interno:
            colunas_traduzidas[coluna_original] = campo_interno

    df = df.rename(columns=colunas_traduzidas)

    # Só mantém colunas que a gente reconhece — evita levar lixo de colunas
    # extras que a pessoa deixou na planilha sem querer.
    colunas_conhecidas = [c for c in df.columns if c in COLUNAS_ACEITAS_NA_IMPORTACAO.values()]
    df = df[colunas_conhecidas]

    # Converte NaN do pandas em None, que o Python/Pydantic entende como "vazio"
    df = df.where(pd.notnull(df), None)

    return df.to_dict(orient="records")